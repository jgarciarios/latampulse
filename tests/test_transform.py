"""
tests/test_transform.py — Test suite de src/transform.py.

Cada test acá corresponde a un bug REAL que encontramos durante el
desarrollo (no son casos hipotéticos) — sirven de regresión: si alguien
en el futuro "simplifica" el código y reintroduce alguno de estos bugs,
un test falla en rojo en vez de descubrirse semanas después con datos
reales corrompidos en silencio.

Corré con: pytest tests/ -v
"""

import json
import datetime

import pandas as pd
import pytest

from src.transform import (
    transform_dolarapi,
    transform_worldbank,
    transform_ibge,
    transform_dane,
    transform_ine_uruguay,
    transform_indec,
    compute_usd_values,
    _try_parse_price_local,
    _try_parse_ine_uy_value,
)


def test_transform_dolarapi_basic(tmp_path):
    envelope = {
        "_metadata": {"extracted_at": "2026-08-05T14:00:00+00:00"},
        "data": [
            {"casa": "oficial", "venta": 1520.0, "compra": 1510.0, "fechaActualizacion": "2026-08-05T13:00:00.000Z"},
            {"casa": "blue", "venta": 1540.0, "compra": 1530.0, "fechaActualizacion": "2026-08-05T13:00:00.000Z"},
        ],
    }
    raw_path = tmp_path / "dolarapi_ar_2026-08-05.json"
    raw_path.write_text(json.dumps(envelope))

    df = transform_dolarapi(raw_path)

    assert len(df) == 2
    assert set(df["rate_type"]) == {"oficial", "blue"}
    assert df[df["rate_type"] == "oficial"]["rate_to_usd"].iloc[0] == 1520.0
    assert (df["country_code"] == "AR").all()


def test_transform_worldbank_ancla_al_ano_no_nulo_de_ppp_factor(tmp_path):
    envelope = {
        "_metadata": {"extracted_at": "2026-08-05T18:13:04+00:00"},
        "data": {
            "gdp_per_capita_ppp_current": [
                {"country": {"id": "AR"}, "date": "2024", "value": None},
                {"country": {"id": "AR"}, "date": "2023", "value": 28251.4},
            ],
            "gdp_per_capita_ppp_constant": [],
            "ppp_conversion_factor": [
                {"country": {"id": "AR"}, "date": "2024", "value": None},
                {"country": {"id": "AR"}, "date": "2023", "value": 305.7},
            ],
        },
    }
    raw_path = tmp_path / "ppp_gdp_2026-08-05.json"
    raw_path.write_text(json.dumps(envelope))

    df = transform_worldbank(raw_path)

    row = df[df["country_code"] == "AR"].iloc[0]
    assert row["year"] == 2023
    assert row["ppp_conversion_factor"] == 305.7
    assert row["gdp_per_capita_ppp_current"] == 28251.4


def test_transform_ibge_usa_d3_para_el_periodo_no_d2(tmp_path):
    envelope = {
        "_metadata": {"extracted_at": "2026-08-05T18:13:41+00:00"},
        "data": [
            {"D2C": "Variável (Código)", "D2N": "Variável", "D3C": "Mês (Código)", "D3N": "Mês", "V": "Valor"},
            {"D2C": "63", "D2N": "IPCA - Variação mensal", "D3C": "202407", "D3N": "julho 2024", "V": "0.38", "MN": "%"},
        ],
    }
    raw_path = tmp_path / "ipca_2026-08-05.json"
    raw_path.write_text(json.dumps(envelope))

    df = transform_ibge(raw_path)

    assert len(df) == 1
    assert df.iloc[0]["period"] == "2024-07-01"
    assert df.iloc[0]["value"] == 0.38


def test_transform_ibge_maneja_valor_no_disponible_de_sidra(tmp_path):
    envelope = {
        "_metadata": {"extracted_at": "2026-08-05T18:13:41+00:00"},
        "data": [
            {"D2C": "x", "D2N": "x", "D3C": "x", "D3N": "x", "V": "x"},
            {"D2C": "63", "D2N": "IPCA", "D3C": "202406", "D3N": "junho 2024", "V": "...", "MN": "%"},
        ],
    }
    raw_path = tmp_path / "ipca_2026-08-05.json"
    raw_path.write_text(json.dumps(envelope))

    df = transform_ibge(raw_path)

    assert pd.isna(df.iloc[0]["value"])


def test_transform_dane_limpia_comillas_de_meses(tmp_path):
    import openpyxl

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "IndicesIPC"
    ws.cell(row=9, column=1, value="Mes")
    ws.cell(row=9, column=2, value=2026)
    ws.cell(row=10, column=1, value="'Enero'")
    ws.cell(row=10, column=2, value=150.5)

    raw_path = tmp_path / "ipc_2026_01.xlsx"
    wb.save(raw_path)

    df = transform_dane(raw_path)

    assert len(df) == 1
    assert df.iloc[0]["period_label"] == "Enero 2026"
    assert df.iloc[0]["value"] == 150.5


def test_ine_uruguay_recupera_ano_de_titulo_de_texto():
    from src.transform import INE_UY_YEAR_TITLE_RE

    match_texto = INE_UY_YEAR_TITLE_RE.search("Año 2019")
    assert match_texto is not None
    assert int(match_texto.group(1)) == 2019


def test_ine_uruguay_integracion_completa_no_hereda_ano_incorrecto(tmp_path):
    df_test = pd.DataFrame([
        [None, 2025, "Enero\xa0", 130.5],
        [None, None, "Febrero\xa0", 131.0],
        [None, "Año 2019", None, None],
        [None, None, "Enero\xa0", 100.0],
    ])

    raw_path = tmp_path / "ipc_test.ods"
    df_test.to_excel(raw_path, engine="odf", header=False, index=False)

    df = transform_ine_uruguay(raw_path)

    fila_2019 = df[df["period"] == "2019-01-01"]
    assert len(fila_2019) == 1, "La fila de 2019 se perdió o quedó mal etiquetada"
    assert fila_2019.iloc[0]["value"] == 100.0

    fila_2025 = df[df["period"] == "2025-01-01"]
    assert fila_2025.iloc[0]["value"] == 130.5


def test_ine_uruguay_value_parser_coma_decimal():
    assert _try_parse_ine_uy_value("99,47") == 99.47
    assert _try_parse_ine_uy_value(99.47) == 99.47


def test_ine_uruguay_value_parser_descarta_texto_no_numerico():
    assert _try_parse_ine_uy_value("Índice de Precios del Consumo") is None
    assert _try_parse_ine_uy_value("Mes") is None


def test_transform_indec_filtra_solo_nacional(tmp_path):
    csv_content = (
        "Codigo;Descripcion;Clasificador;Periodo;Indice_IPC;v_m_IPC;v_i_a_IPC;Region\n"
        "0;Nivel General;Nivel General;202606;3450,77;2,1;45,3;Nacional\n"
        "0;Nivel General;Nivel General;202606;99,1;NA;NA;GBA\n"
    )
    raw_path = tmp_path / "ipc_2026-08-05.csv"
    raw_path.write_text(csv_content, encoding="latin-1")

    df = transform_indec(raw_path)

    assert len(df) > 0
    assert "GBA" not in df["period_label"].str.cat()


def test_transform_indec_convierte_coma_decimal(tmp_path):
    csv_content = (
        "Codigo;Descripcion;Clasificador;Periodo;Indice_IPC;v_m_IPC;v_i_a_IPC;Region\n"
        "0;Nivel General;Nivel General;202606;3450,77;2,1;45,3;Nacional\n"
    )
    raw_path = tmp_path / "ipc_2026-08-05.csv"
    raw_path.write_text(csv_content, encoding="latin-1")

    df = transform_indec(raw_path)

    indice_row = df[df["indicator"] == "indec_0_indice"]
    assert indice_row.iloc[0]["value"] == pytest.approx(3450.77)


def test_compute_usd_values_elimina_filas_con_price_local_invalido():
    prices = pd.DataFrame([
        {"country": "AR", "item_name": "Netflix", "price_local": 14999, "currency": "ARS", "date_captured": "2026-08-09"},
        {"country": "BR", "item_name": "Café roto", "price_local": datetime.datetime(2026, 5, 1), "currency": "BRL", "date_captured": "2026-08-09"},
    ])
    fx = pd.DataFrame([{"country_code": "AR", "date": "2026-08-05", "rate_type": "oficial", "rate_to_usd": 1520.0}])
    ppp = pd.DataFrame([{"country_code": "AR", "year": 2024, "ppp_conversion_factor": 419.29}])

    result = compute_usd_values(prices, fx, ppp)

    assert len(result) == 1
    assert result.iloc[0]["item_name"] == "Netflix"


def test_compute_usd_values_nominal_none_sin_exchange_rate():
    prices = pd.DataFrame([
        {"country": "BR", "item_name": "Netflix", "price_local": 23.90, "currency": "BRL", "date_captured": "2026-08-09"},
    ])
    fx = pd.DataFrame([{"country_code": "AR", "date": "2026-08-05", "rate_type": "oficial", "rate_to_usd": 1520.0}])
    ppp = pd.DataFrame([{"country_code": "BR", "year": 2024, "ppp_conversion_factor": 2.48}])

    result = compute_usd_values(prices, fx, ppp)

    assert pd.isna(result.iloc[0]["price_usd_nominal"])
    assert result.iloc[0]["price_usd_ppp"] == pytest.approx(23.90 / 2.48)


def test_try_parse_price_local_coma_y_punto():
    assert _try_parse_price_local("23,90") == 23.90
    assert _try_parse_price_local("23.90") == 23.90
    assert _try_parse_price_local(23.90) == 23.90
    assert _try_parse_price_local(datetime.datetime(2026, 5, 1)) is None
