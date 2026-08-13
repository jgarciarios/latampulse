"""
transform.py — Normalización de datos crudos hacia el schema de LatamPulse.

Regla de diseño: acá SÍ transformamos. A diferencia de extract.py (que
guarda todo tal cual vino), este archivo lee de data/raw/, limpia,
normaliza formatos heterogéneos (JSON, Excel, ODS, CSV) hacia una
estructura común, y calcula los campos derivados (usd_nominal, usd_ppp).
"""

import json
import re
from pathlib import Path

import pandas as pd

from src.utils import get_logger

logger = get_logger("latampulse.transform")

RAW_DATA_DIR = Path(__file__).resolve().parent.parent / "data" / "raw"


def load_latest_raw(source: str, extension: str = "json") -> Path:
    source_dir = RAW_DATA_DIR / source
    if not source_dir.exists():
        raise FileNotFoundError(f"No existe data/raw/{source}/ — ¿corriste extract.py?")
    matching_files = sorted(source_dir.glob(f"*.{extension}"))
    if not matching_files:
        raise FileNotFoundError(f"No hay archivos .{extension} en data/raw/{source}/")
    return matching_files[-1]


def transform_dolarapi(raw_path: Path) -> pd.DataFrame:
    """
    Convierte el JSON de dolarapi a filas de exchange_rates.
    Usamos 'venta' como rate_to_usd (referencia estándar en Argentina
    para cotizar "el dólar está a $X").
    """
    with open(raw_path, "r", encoding="utf-8") as f:
        envelope = json.load(f)
    quotes = envelope["data"]
    fetched_at = envelope["_metadata"]["extracted_at"]
    rows = []
    for quote in quotes:
        rows.append({
            "country_code": "AR",
            "date": quote["fechaActualizacion"][:10],
            "rate_type": quote["casa"],
            "rate_to_usd": quote["venta"],
            "source": "dolarapi",
            "fetched_at": fetched_at,
        })
    df = pd.DataFrame(rows)
    logger.info(f"transform_dolarapi: {len(df)} filas generadas desde {raw_path.name}")
    return df


def transform_worldbank(raw_path: Path) -> pd.DataFrame:
    """
    Convierte el JSON de World Bank a filas de ppp_factors.

    Lógica de "año más reciente no-nulo": el Banco Mundial no actualiza
    todos los indicadores todos los años al mismo tiempo (ej. PPP de 2024
    puede venir null todavía mientras GDP nominal de 2024 ya está). Por
    eso NO asumimos que todos los indicadores comparten el mismo año de
    referencia para un país — cada uno se resuelve independientemente a
    su propio año más reciente con dato.

    Año de referencia de la fila: se ancla al año más reciente no-nulo de
    'ppp_conversion_factor' (el indicador central del proyecto). Los
    valores de GDP per cápita PPP se toman para ESE MISMO año si existen;
    si ese indicador no tiene dato para ese año puntual, queda None en
    vez de rellenar con un año distinto — mezclar años distintos en la
    misma fila sería más confuso que mostrar "sin dato".
    """
    with open(raw_path, "r", encoding="utf-8") as f:
        envelope = json.load(f)

    data = envelope["data"]
    fetched_at = envelope["_metadata"]["extracted_at"]

    def latest_nonnull_by_country(records: list) -> dict:
        best = {}
        for r in records:
            country_code = r["country"]["id"]
            year = int(r["date"])
            value = r["value"]
            if value is None:
                continue
            if country_code not in best or year > best[country_code][0]:
                best[country_code] = (year, value)
        return best

    def all_values_by_country_year(records: list) -> dict:
        result = {}
        for r in records:
            country_code = r["country"]["id"]
            year = int(r["date"])
            value = r["value"]
            if value is None:
                continue
            result.setdefault(country_code, {})[year] = value
        return result

    ppp_factor_latest = latest_nonnull_by_country(data.get("ppp_conversion_factor", []))
    gdp_ppp_current_by_year = all_values_by_country_year(data.get("gdp_per_capita_ppp_current", []))
    gdp_ppp_constant_by_year = all_values_by_country_year(data.get("gdp_per_capita_ppp_constant", []))

    all_countries = (
        set(ppp_factor_latest)
        | set(gdp_ppp_current_by_year)
        | set(gdp_ppp_constant_by_year)
    )

    rows = []
    for country_code in sorted(all_countries):
        if country_code in ppp_factor_latest:
            year, ppp_factor = ppp_factor_latest[country_code]
        else:
            candidate_years = list(gdp_ppp_current_by_year.get(country_code, {}).keys())
            candidate_years += list(gdp_ppp_constant_by_year.get(country_code, {}).keys())
            year = max(candidate_years) if candidate_years else None
            ppp_factor = None
            logger.warning(f"Sin ppp_conversion_factor para {country_code} — usando año {year} de otro indicador")

        gdp_ppp_current = gdp_ppp_current_by_year.get(country_code, {}).get(year)
        gdp_ppp_constant = gdp_ppp_constant_by_year.get(country_code, {}).get(year)

        rows.append({
            "country_code": country_code,
            "year": year,
            "ppp_conversion_factor": ppp_factor,
            "gdp_per_capita_ppp_current": gdp_ppp_current,
            "gdp_per_capita_ppp_constant": gdp_ppp_constant,
            "source": "world_bank",
            "fetched_at": fetched_at,
        })

    df = pd.DataFrame(rows)
    logger.info(f"transform_worldbank: {len(df)} filas generadas desde {raw_path.name}")
    return df


def transform_ibge(raw_path: Path) -> pd.DataFrame:
    """
    Convierte el JSON de SIDRA (IPCA Brasil) a filas de inflation_indices.

    Estructura real de apisidra: la respuesta es una lista donde el
    elemento [0] es un diccionario de ETIQUETAS de columna (no datos —
    ej. {"V": "Valor", "D2C": "Mês (Código)", ...}), y los elementos
    [1:] son las filas de datos reales usando esas mismas claves.
    Saltar el elemento [0] es obligatorio, no opcional.

    SIDRA marca los valores no disponibles con el string literal "...",
    no con null — hay que filtrarlo explícitamente o termina como texto
    en una columna que debería ser numérica.
    """
    with open(raw_path, "r", encoding="utf-8") as f:
        envelope = json.load(f)

    payload = envelope["data"]
    fetched_at = envelope["_metadata"]["extracted_at"]

    if len(payload) < 2:
        logger.warning(f"transform_ibge: {raw_path.name} no tiene filas de datos, solo header")
        return pd.DataFrame()

    data_rows = payload[1:]

    rows = []
    for r in data_rows:
        raw_value = r.get("V")
        if raw_value in (None, "...", "-", ""):
            value = None
        else:
            value = float(raw_value)

        period_code = r.get("D3C")  # formato YYYYMM, ej. "202407" — confirmado
        # contra el JSON real: en esta tabla/variable de SIDRA, D2C/D2N
        # identifican la VARIABLE (código 63 = "IPCA - Variação mensal"),
        # y D3C/D3N identifican el MES. No es un layout universal de SIDRA
        # — cambia según qué combinación de tabla/variable se pida — así
        # que si en el futuro cambiamos IPCA_TABLE o la variable en
        # extract.py, hay que re-verificar esto con el mismo one-liner de
        # inspección, no asumir que D3C sigue siendo el mes.
        period_date = None
        if period_code and len(period_code) == 6:
            period_date = f"{period_code[:4]}-{period_code[4:6]}-01"

        rows.append({
            "country_code": "BR",
            "period": period_date,
            "period_label": r.get("D3N"),  # ej. "julho 2024" — ver nota arriba sobre D2 vs D3
            "indicator": "ipca_variacao_mensal",
            "value": value,
            "unit": r.get("MN"),
            "source": "ibge_sidra",
            "fetched_at": fetched_at,
        })

    df = pd.DataFrame(rows)
    logger.info(f"transform_ibge: {len(df)} filas generadas desde {raw_path.name}")
    return df


DANE_MONTH_NAME_TO_NUMBER = {
    "enero": 1, "febrero": 2, "marzo": 3, "abril": 4, "mayo": 5, "junio": 6,
    "julio": 7, "agosto": 8, "septiembre": 9, "octubre": 10, "noviembre": 11, "diciembre": 12,
}


def transform_dane(raw_path: Path) -> pd.DataFrame:
    """
    Parsea el Excel de DANE (hoja 'IndicesIPC') a filas de inflation_indices.

    Estructura confirmada por inspección directa del archivo real:
      - Fila 9: header. Columna A = "Mes", columnas B-Y = años (2003-2026
        aprox.), columnas Z/AA/AB vacías (se ignoran).
      - Filas 10-21: un mes por fila (Enero a Diciembre), valores del
        índice como float nativo (no hace falta convertir de texto).
      - Los nombres de mes vienen con comillas simples embebidas en el
        string literal (ej. "\'Enero\'"), hay que hacer .strip("\'").
      - El año en curso tiene columnas con None a partir del mes que
        todavía no se publicó — se omiten esas celdas, no se rellenan.

    Sobre fetched_at: a diferencia de los JSON (que llevan _metadata con
    extracted_at), este archivo no tiene metadata de extracción embebida
    — el nombre codifica el PERÍODO del dato (ipc_2026_06.xlsx = datos
    hasta junio 2026), no cuándo lo bajamos. Usamos la fecha de
    modificación del archivo en disco como aproximación razonable.
    """
    import openpyxl

    wb = openpyxl.load_workbook(raw_path, data_only=True)
    ws = wb["IndicesIPC"]

    HEADER_ROW = 9
    DATA_START_ROW = 10
    DATA_END_ROW = 21
    YEAR_COL_START = 2
    YEAR_COL_END = 25

    mtime = raw_path.stat().st_mtime
    fetched_at = pd.Timestamp.fromtimestamp(mtime, tz="UTC").isoformat()

    years_by_col = {}
    for col_idx in range(YEAR_COL_START, YEAR_COL_END + 1):
        val = ws.cell(row=HEADER_ROW, column=col_idx).value
        if val is not None:
            years_by_col[col_idx] = int(val)

    rows = []
    for row_idx in range(DATA_START_ROW, DATA_END_ROW + 1):
        month_raw = ws.cell(row=row_idx, column=1).value
        if month_raw is None:
            continue

        month_name = str(month_raw).strip().strip("'").strip()
        month_number = DANE_MONTH_NAME_TO_NUMBER.get(month_name.lower())
        if month_number is None:
            logger.warning(f"transform_dane: mes no reconocido \'{month_name}\' en fila {row_idx}, se omite")
            continue

        for col_idx, year in years_by_col.items():
            value = ws.cell(row=row_idx, column=col_idx).value
            if value is None:
                continue

            period_date = f"{year:04d}-{month_number:02d}-01"
            rows.append({
                "country_code": "CO",
                "period": period_date,
                "period_label": f"{month_name} {year}",
                "indicator": "ipc_indice",
                "value": float(value),
                "unit": None,
                "source": "dane",
                "fetched_at": fetched_at,
            })

    df = pd.DataFrame(rows)
    logger.info(f"transform_dane: {len(df)} filas generadas desde {raw_path.name}")
    return df

INE_UY_MONTH_NAME_TO_NUMBER = {
    "enero": 1, "febrero": 2, "marzo": 3, "abril": 4, "mayo": 5, "junio": 6,
    "julio": 7, "agosto": 8, "septiembre": 9, "setiembre": 9,
    "octubre": 10, "noviembre": 11, "diciembre": 12,
}

INE_UY_YEAR_TITLE_RE = re.compile(r"a[\u00f1n]o\s*(\d{4})", re.IGNORECASE)


def _try_parse_ine_uy_value(raw_value):
    """
    Convierte el valor crudo de la columna 3 a float, manejando dos
    casos reales encontrados en el archivo:
      1. Números con coma decimal ("99,47" en vez de "99.47").
      2. Texto que no es un número (headers/metadata colados, ej.
         "Índice de Precios del Consumo") — devuelve None en vez de
         levantar ValueError.
    """
    if isinstance(raw_value, (int, float)):
        return float(raw_value)
    if isinstance(raw_value, str):
        cleaned = raw_value.strip().replace(",", ".")
        try:
            return float(cleaned)
        except ValueError:
            return None
    return None


def transform_ine_uruguay(raw_path: Path) -> pd.DataFrame:
    """
    Parsea el .ods de DGI/INE (serie IPC Uruguay) a filas de inflation_indices.

    Bugs reales encontrados y corregidos en esta función:
      1. ffill() vectorizado sobre año de tipo mixto propagaba el año
         equivocado cuando solo aparecía como texto de título ("Año
         2019") — resuelto parseando el año fila por fila.
      2. Valores con coma decimal ("99,47") y filas de header con texto
         no-numérico en la columna de valor — resueltos con
         _try_parse_ine_uy_value(), que normaliza la coma y devuelve
         None para saltear filas de ruido antes del lookup de mes.
    """
    df_raw = pd.read_excel(raw_path, sheet_name=0, engine="odf", header=None)

    mtime = raw_path.stat().st_mtime
    fetched_at = pd.Timestamp.fromtimestamp(mtime, tz="UTC").isoformat()

    current_year = None
    rows = []

    for _, r in df_raw.iterrows():
        col_year_raw, col_month_raw, col_value_raw = r[1], r[2], r[3]

        year_found = None
        if pd.notna(col_year_raw):
            if isinstance(col_year_raw, (int, float)):
                try:
                    year_found = int(col_year_raw)
                except (ValueError, TypeError):
                    pass
            elif isinstance(col_year_raw, str):
                stripped = col_year_raw.strip()
                if stripped.isdigit() and len(stripped) == 4:
                    year_found = int(stripped)
                else:
                    match = INE_UY_YEAR_TITLE_RE.search(stripped)
                    if match:
                        year_found = int(match.group(1))
        if year_found is not None:
            current_year = year_found

        if pd.isna(col_month_raw) or pd.isna(col_value_raw):
            continue

        value = _try_parse_ine_uy_value(col_value_raw)
        if value is None:
            continue

        month_name = str(col_month_raw).replace("\xa0", "").strip()
        month_number = INE_UY_MONTH_NAME_TO_NUMBER.get(month_name.lower())
        if month_number is None:
            logger.warning(f"transform_ine_uruguay: mes no reconocido '{month_name}', se omite")
            continue

        if current_year is None:
            logger.warning(
                f"transform_ine_uruguay: fila con mes '{month_name}' sin año "
                f"de referencia conocido todavía — se omite"
            )
            continue

        period_date = f"{current_year:04d}-{month_number:02d}-01"
        rows.append({
            "country_code": "UY",
            "period": period_date,
            "period_label": f"{month_name} {current_year}",
            "indicator": "ipc_indice",
            "value": value,
            "unit": None,
            "source": "ine_uy_dgi",
            "fetched_at": fetched_at,
        })

    df = pd.DataFrame(rows)
    logger.info(f"transform_ine_uruguay: {len(df)} filas generadas desde {raw_path.name}")
    return df

def transform_indec(raw_path: Path) -> pd.DataFrame:
    """
    Parsea el CSV de INDEC a filas de inflation_indices para Argentina.
    Separador ';', encoding latin-1. Filtra Region == 'Nacional'.
    Descompone Indice_IPC / v_m_IPC / v_i_a_IPC en filas separadas
    (formato long, consistente con los otros 3 países).
    """
    df = pd.read_csv(raw_path, sep=";", encoding="latin-1")
    df = df[df["Region"] == "Nacional"].copy()

    mtime = raw_path.stat().st_mtime
    fetched_at = pd.Timestamp.fromtimestamp(mtime, tz="UTC").isoformat()

    def parse_indec_value(raw):
        if pd.isna(raw):
            return None
        if isinstance(raw, (int, float)):
            return float(raw)
        s = str(raw).strip()
        if s == "" or s.upper() == "NA":
            return None
        s = s.replace(",", ".")
        try:
            return float(s)
        except ValueError:
            return None

    metric_columns = {
        "indice": ("Indice_IPC", "indice"),
        "variacion_mensual": ("v_m_IPC", "%"),
        "variacion_interanual": ("v_i_a_IPC", "%"),
    }

    rows = []
    for _, r in df.iterrows():
        periodo_str = str(int(r["Periodo"]))
        if len(periodo_str) != 6:
            logger.warning(f"transform_indec: Periodo con formato inesperado \'{periodo_str}\', se omite fila")
            continue

        period_date = f"{periodo_str[:4]}-{periodo_str[4:]}-01"
        codigo = r["Codigo"]
        descripcion = r["Descripcion"]

        for metric_name, (col_name, unit) in metric_columns.items():
            value = parse_indec_value(r[col_name])
            if value is None:
                continue

            rows.append({
                "country_code": "AR",
                "period": period_date,
                "period_label": f"{descripcion} ({periodo_str})",
                "indicator": f"indec_{codigo}_{metric_name}",
                "value": value,
                "unit": unit,
                "source": "indec",
                "fetched_at": fetched_at,
            })

    result_df = pd.DataFrame(rows)
    logger.info(f"transform_indec: {len(result_df)} filas generadas desde {raw_path.name}")
    return result_df

def _try_parse_price_local(raw_value):
    """
    Convierte price_local a float, manejando dos problemas reales
    encontrados en research_template_1.xlsx:
      1. Valores con coma decimal cargados como texto en Excel.
      2. Celdas donde quedó una fecha en vez de un precio (Excel las
         guarda como datetime.datetime) — se devuelve None para que el
         llamador saltee la fila con un warning, no se adivina el precio.
    """
    if isinstance(raw_value, bool):
        return None
    if isinstance(raw_value, (int, float)):
        return float(raw_value)
    if isinstance(raw_value, str):
        cleaned = raw_value.strip().replace(",", ".")
        try:
            return float(cleaned)
        except ValueError:
            return None
    return None


def compute_usd_values(prices_df: pd.DataFrame, exchange_rates_df: pd.DataFrame, ppp_factors_df: pd.DataFrame) -> pd.DataFrame:
    """
    Agrega price_usd_nominal y price_usd_ppp a un DataFrame de precios.
    Filas con price_local no numérico se ELIMINAN del resultado (con
    warning), nunca se inventa el precio real.
    """
    df = prices_df.copy()
    df["price_usd_nominal"] = None
    df["price_usd_ppp"] = None

    fx_oficial = exchange_rates_df[exchange_rates_df["rate_type"] == "oficial"].copy()
    fx_oficial["date"] = pd.to_datetime(fx_oficial["date"])

    ppp = ppp_factors_df.copy()

    countries_sin_fx = set(df["country"].unique()) - set(fx_oficial["country_code"].unique())
    if countries_sin_fx:
        logger.warning(
            f"Sin exchange_rates 'oficial' para: {sorted(countries_sin_fx)} — "
            f"price_usd_nominal va a quedar sin dato para esos países"
        )

    filas_invalidas = []
    for idx, row in df.iterrows():
        country = row["country"]
        price_local = _try_parse_price_local(row["price_local"])
        date_captured = pd.to_datetime(row["date_captured"])

        if price_local is None:
            item_name = row.get("item_name", "?")
            logger.warning(
                f"compute_usd_values: price_local inválido para "
                f"'{item_name}' ({country}), fila índice {idx} — valor "
                f"crudo: {row['price_local']!r}. Fila eliminada del "
                f"resultado, revisar el Excel manualmente."
            )
            filas_invalidas.append(idx)
            continue

        df.at[idx, "price_local"] = price_local

        country_fx = fx_oficial[fx_oficial["country_code"] == country]
        if not country_fx.empty:
            closest_idx = (country_fx["date"] - date_captured).abs().idxmin()
            rate = country_fx.loc[closest_idx, "rate_to_usd"]
            if pd.notna(rate) and rate != 0:
                df.at[idx, "price_usd_nominal"] = price_local / rate

        country_ppp = ppp[ppp["country_code"] == country]
        if not country_ppp.empty:
            year_captured = date_captured.year
            closest_idx = (country_ppp["year"] - year_captured).abs().idxmin()
            factor = country_ppp.loc[closest_idx, "ppp_conversion_factor"]
            if pd.notna(factor) and factor != 0:
                df.at[idx, "price_usd_ppp"] = price_local / factor

    filas_salteadas = len(filas_invalidas)
    if filas_invalidas:
        df = df.drop(index=filas_invalidas).reset_index(drop=True)

    logger.info(
        f"compute_usd_values: {len(df)} filas procesadas ({filas_salteadas} "
        f"salteadas por price_local inválido) — "
        f"{df['price_usd_nominal'].notna().sum()} con USD nominal, "
        f"{df['price_usd_ppp'].notna().sum()} con USD PPP"
    )
    return df
